# imports section
# ---------------- Core ----------------
import os
import datetime
from typing import List, Optional
import time
import chromadb

# FastAPI & Server
from fastapi import FastAPI, UploadFile, File, HTTPException, Depends
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware

# Google Gemini API
import google.generativeai as genai

# ---------------- Document Handling ----------------
import pdfplumber             # Extract text from PDFs
import docx                   # Read .docx Word files
from PIL import Image         # Pillow, needed for OCR
import pytesseract            # OCR for images
import openpyxl               # Read Excel files

# ---------------- Database & Utilities ----------------
from sqlalchemy.orm import Session
from dotenv import load_dotenv
from fastapi_utils.tasks import repeat_every





# api setup area
app = FastAPI()

genai.configure(api_key="AIzaSyA20oTijrHh6sip6JWxdvke0ZvrVG02HsA")
chroma_client = chromadb.PersistentClient(path="./chroma_db")
collection = chroma_client.get_or_create_collection("documents")



# documents upload area
# ---------------- Document Extraction Helpers ----------------

def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from PDF using pdfplumber"""
    text = ""
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text() or ""
    return text.strip()


def extract_text_from_docx(file_path: str) -> str:
    """Extract text from Word .docx file"""
    import docx
    doc = docx.Document(file_path)
    return "\n".join([para.text for para in doc.paragraphs]).strip()


def extract_text_from_txt(file_path: str) -> str:
    """Read text from plain .txt file"""
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read().strip()


def extract_text_from_image(file_path: str) -> str:
    """Extract text from image using pytesseract"""
    img = Image.open(file_path)
    return pytesseract.image_to_string(img).strip()


def extract_text_from_excel(file_path: str) -> str:
    """Extract text from Excel file (.xls, .xlsx)"""
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    text = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            row_text = [str(cell) for cell in row if cell is not None]
            if row_text:
                text.append(" ".join(row_text))
    return "\n".join(text).strip()


# ---------------- Dispatcher ----------------

def extract_text(file_path: str, filename: str) -> str:
    """Choose extractor based on file extension"""
    ext = filename.lower().split(".")[-1]

    if ext == "pdf":
        return extract_text_from_pdf(file_path)
    elif ext == "docx":
        return extract_text_from_docx(file_path)
    elif ext == "txt":
        return extract_text_from_txt(file_path)
    elif ext in ["jpg", "jpeg", "png"]:
        return extract_text_from_image(file_path)
    elif ext in ["xls", "xlsx"]:
        return extract_text_from_excel(file_path)
    else:
        raise ValueError(f"Unsupported file type: .{ext}")



from fastapi import BackgroundTasks

@app.post("/upload/")
async def upload_file(file: UploadFile = File(...)):
    # Save uploaded file temporarily
    temp_file_path = f"temp_{int(time.time())}_{file.filename}"
    with open(temp_file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)

    try:
        # Extract text from file
        text = extract_text(temp_file_path, file.filename)
    except Exception as e:
        os.remove(temp_file_path)
        raise HTTPException(status_code=400, detail=str(e))

    # Split text into 500-character chunks
    chunks = [text[i:i+500] for i in range(0, len(text), 500)]

    # Current timestamp
    uploaded_at = int(time.time())

    # Add each chunk to chroma collection
    for idx, chunk in enumerate(chunks):
        collection.add(
            documents=[chunk],
            metadatas=[{"filename": file.filename, "uploaded_at": uploaded_at}],
            ids=[f"{file.filename}_{uploaded_at}_{idx}"]
        )

    # Delete temp file
    os.remove(temp_file_path)

    return JSONResponse(content={"message": f"Stored {len(chunks)} chunks from {file.filename}."})


def cleanup_old_docs():
    now = int(time.time())
    cutoff = now - 48 * 3600  # 48 hours ago

    # Get all documents
    all_docs = collection.get(include=["ids", "metadatas"])

    ids_to_delete = []
    for doc_id, metadata in zip(all_docs["ids"], all_docs["metadatas"]):
        uploaded_at = metadata.get("uploaded_at", 0)
        if uploaded_at < cutoff:
            ids_to_delete.append(doc_id)

    if ids_to_delete:
        collection.delete(ids=ids_to_delete)


@app.on_event("startup")
@repeat_every(seconds=3600)  # Run every hour
def periodic_cleanup():
    cleanup_old_docs()




# user input





# content finding from database

@app.post("/query/")
async def query_documents(question: str):
    # Step 1: Search in ChromaDB
    results = collection.query(query_texts=[question], n_results=3, include=["documents"])
    documents = results.get("documents", [[]])[0]

    if not documents:
        return JSONResponse(content={"answer": "No relevant documents found."})

    # Step 2: Build context
    context = "\n".join(documents)

    # Step 3: Call Gemini
    model = genai.GenerativeModel("gemini-1.5-flash")
    response = model.generate_content(
        f"Answer the following question based only on this context:\n\n{context}\n\nQuestion: {question}"
    )

    return JSONResponse(content={
        "answer": response.text,
        "context": context
    })













# Serve frontend (index.html, script.js, style.css)
from fastapi.staticfiles import StaticFiles

# Serve frontend (index.html, script.js, style.css)
app.mount("/", StaticFiles(directory="../frontend", html=True), name="frontend")